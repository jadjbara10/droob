"""Extract بيانات المجمعات from Excel → JSON for the TS seed script"""
import openpyxl, json, os

EXCEL_PATH = r"d:\trans_app\المجمعات_بالاردن.xlsx"
OUTPUT_RAW = os.path.join(os.path.dirname(__file__), "excel-raw-data.json")
OUTPUT_JSON = os.path.join(os.path.dirname(__file__), "..", "..", "backend", "src", "data", "jordan-terminals.json")

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

print(f"📊 Excel: {ws.max_row} rows × {ws.max_column} columns")
print(f"   Sheet name: {ws.title}")
print()

# Print first 10 rows to understand structure
print("First 10 rows:")
for r in range(1, min(ws.max_row + 1, 11)):
    values = [str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)]
    print(f"  Row {r}: {values}")

# Extract all rows as list of lists for the TS extract-excel-stops.ts script
all_rows = []
for r in range(1, ws.max_row + 1):
    row_values = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    all_rows.append(row_values)

# Write the raw data JSON for the TS script to consume
os.makedirs(os.path.dirname(OUTPUT_RAW), exist_ok=True)
with open(OUTPUT_RAW, "w", encoding="utf-8") as f:
    json.dump(all_rows, f, ensure_ascii=False, indent=2)

print(f"\n✅ Raw data saved: {OUTPUT_RAW}")
print(f"   Total rows: {len(all_rows)}")
print(f"   Non-empty governorate rows: {sum(1 for r in all_rows if r[0])}")